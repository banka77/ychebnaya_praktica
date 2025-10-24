from fastapi import FastAPI, HTTPException
from fastapi.responses import HTMLResponse, Response, FileResponse
from pydantic import BaseModel
import psycopg2
from psycopg2.extras import RealDictCursor
from typing import List, Dict, Optional
import os
import pandas as pd
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import io
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

app = FastAPI()

# Настройки подключения к БД
DB_CONFIG = {
    'host': '10.115.0.67',
    'database': 'edu_practice01_03',
    'user': 'user_03',
    'password': 'password3'
}


def get_db_connection():
    conn = psycopg2.connect(**DB_CONFIG)
    return conn


# Pydantic модели для валидации
class CarCreate(BaseModel):
    model_id: int
    color: str
    price: float
    year: int
    engine_type: str
    transmission: str
    mileage: int
    status: str


class ManufacturerCreate(BaseModel):
    name: str
    country: str
    year_founded: int


class ModelCreate(BaseModel):
    name: str
    manufacturer_id: int
    year: int
    body_type: str


# Регистрируем шрифт для кириллицы в PDF
try:
    pdfmetrics.registerFont(TTFont('Arial', 'arial.ttf'))
except:
    # Если шрифт не найден, используем стандартный
    pass


# Читаем HTML файл
def get_html():
    html_path = os.path.join(os.path.dirname(__file__), 'index.html')
    with open(html_path, 'r', encoding='utf-8') as f:
        return f.read()


@app.get("/", response_class=HTMLResponse)
async def index():
    return get_html()


# ============ CARS API ============
@app.get("/api/cars")
async def get_cars():
    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT c.id, man.name as manufacturer, m.name as model, c.color, 
                   c.price, c.year, c.engine_type, c.transmission, c.mileage, c.status,
                   c.model_id
            FROM cars c
            JOIN models m ON c.model_id = m.id
            JOIN manufacturers man ON m.manufacturer_id = man.id
            ORDER BY c.id
        """)
        cars = cur.fetchall()
        cur.close()
        conn.close()
        return cars
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/cars")
async def create_car(car: CarCreate):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO cars (model_id, color, price, year, engine_type, transmission, mileage, status)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (car.model_id, car.color, car.price, car.year, car.engine_type,
              car.transmission, car.mileage, car.status))
        car_id = cur.fetchone()[0]
        conn.commit()
        cur.close()
        conn.close()
        return {'success': True, 'id': car_id}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/cars/{car_id}")
async def update_car(car_id: int, car: CarCreate):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            UPDATE cars 
            SET model_id=%s, color=%s, price=%s, year=%s, engine_type=%s, 
                transmission=%s, mileage=%s, status=%s
            WHERE id=%s
        """, (car.model_id, car.color, car.price, car.year, car.engine_type,
              car.transmission, car.mileage, car.status, car_id))
        conn.commit()
        cur.close()
        conn.close()
        return {'success': True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/cars/{car_id}")
async def delete_car(car_id: int):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM cars WHERE id = %s", (car_id,))
        conn.commit()
        cur.close()
        conn.close()
        return {'success': True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============ MANUFACTURERS API ============
@app.get("/api/manufacturers")
async def get_manufacturers():
    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT * FROM manufacturers ORDER BY id")
        manufacturers = cur.fetchall()
        cur.close()
        conn.close()
        return manufacturers
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/manufacturers")
async def create_manufacturer(manufacturer: ManufacturerCreate):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO manufacturers (name, country, year_founded)
            VALUES (%s, %s, %s)
            RETURNING id
        """, (manufacturer.name, manufacturer.country, manufacturer.year_founded))
        man_id = cur.fetchone()[0]
        conn.commit()
        cur.close()
        conn.close()
        return {'success': True, 'id': man_id}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/manufacturers/{man_id}")
async def update_manufacturer(man_id: int, manufacturer: ManufacturerCreate):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            UPDATE manufacturers 
            SET name=%s, country=%s, year_founded=%s
            WHERE id=%s
        """, (manufacturer.name, manufacturer.country, manufacturer.year_founded, man_id))
        conn.commit()
        cur.close()
        conn.close()
        return {'success': True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/manufacturers/{man_id}")
async def delete_manufacturer(man_id: int):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM manufacturers WHERE id = %s", (man_id,))
        conn.commit()
        cur.close()
        conn.close()
        return {'success': True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============ MODELS API ============
@app.get("/api/models")
async def get_models():
    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT m.id, m.name, man.name as manufacturer, m.year, m.body_type,
                   m.manufacturer_id
            FROM models m
            JOIN manufacturers man ON m.manufacturer_id = man.id
            ORDER BY m.id
        """)
        models = cur.fetchall()
        cur.close()
        conn.close()
        return models
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/models")
async def create_model(model: ModelCreate):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO models (name, manufacturer_id, year, body_type)
            VALUES (%s, %s, %s, %s)
            RETURNING id
        """, (model.name, model.manufacturer_id, model.year, model.body_type))
        model_id = cur.fetchone()[0]
        conn.commit()
        cur.close()
        conn.close()
        return {'success': True, 'id': model_id}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/models/{model_id}")
async def update_model(model_id: int, model: ModelCreate):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            UPDATE models 
            SET name=%s, manufacturer_id=%s, year=%s, body_type=%s
            WHERE id=%s
        """, (model.name, model.manufacturer_id, model.year, model.body_type, model_id))
        conn.commit()
        cur.close()
        conn.close()
        return {'success': True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/models/{model_id}")
async def delete_model(model_id: int):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM models WHERE id = %s", (model_id,))
        conn.commit()
        cur.close()
        conn.close()
        return {'success': True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============ EXPORT API ============
@app.get("/api/export/contract/{car_id}")
async def export_contract(car_id: int):
    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        # Получаем данные автомобиля
        cur.execute("""
            SELECT c.*, m.name as model_name, man.name as manufacturer_name,
                   man.country as manufacturer_country
            FROM cars c
            JOIN models m ON c.model_id = m.id
            JOIN manufacturers man ON m.manufacturer_id = man.id
            WHERE c.id = %s
        """, (car_id,))
        car = cur.fetchone()

        cur.close()
        conn.close()

        if not car:
            raise HTTPException(status_code=404, detail="Автомобиль не найден")

        # Создаем Word-документ вручную (HTML с расширением .doc)
        contract_content = f"""<html>
<head>
<meta charset="UTF-8">
<title>Договор купли-продажи автомобиля</title>
<style>
body {{ font-family: Arial, sans-serif; margin: 40px; line-height: 1.6; }}
.header {{ text-align: center; font-weight: bold; font-size: 16pt; margin-bottom: 30px; }}
.section {{ margin-bottom: 20px; }}
.section-title {{ font-weight: bold; font-size: 12pt; margin-bottom: 10px; }}
.signature {{ margin-top: 50px; }}
.signature-line {{ margin-top: 40px; }}
</style>
</head>
<body>
<div class="header">ДОГОВОР КУПЛИ-ПРОДАЖИ АВТОМОБИЛЯ</div>
<div style="text-align: center; margin-bottom: 30px;">г. Москва {datetime.datetime.now().strftime('%d.%m.%Y')} г.</div>

<div class="section">
    <div class="section-title">1. Стороны договора</div>
    <p><strong>Продавец:</strong> Автосалон "АвтоМир"</p>
    <p><strong>Покупатель:</strong> [ФИО покупателя]</p>
</div>

<div class="section">
    <div class="section-title">2. Предмет договора</div>
    <p>2.1. Продавец обязуется передать, а Покупатель - принять и оплатить автомобиль с характеристиками:</p>
    <ul>
        <li><strong>Марка, модель:</strong> {car['manufacturer_name']} {car['model_name']}</li>
        <li><strong>Год выпуска:</strong> {car['year']}</li>
        <li><strong>Цвет:</strong> {car['color']}</li>
        <li><strong>Пробег:</strong> {car['mileage']} км</li>
        <li><strong>Цена:</strong> {car['price']:,.2f} руб.</li>
        <li><strong>Тип двигателя:</strong> {car['engine_type']}</li>
        <li><strong>Коробка передач:</strong> {car['transmission']}</li>
        <li><strong>VIN:</strong> VIN_{car['id']}</li>
        <li><strong>Двигатель №:</strong> ДВ_{car['id']}</li>
        <li><strong>Кузов №:</strong> КУ_{car['id']}</li>
    </ul>
</div>

<div class="section">
    <div class="section-title">3. Цена и порядок оплаты</div>
    <p>3.1. Общая цена: <strong>{car['price']:,.2f} руб.</strong></p>
</div>

<div class="section">
    <div class="section-title">4. Условия передачи автомобиля</div>
    <p>4.1. Автомобиль передается Покупателю в день подписания договора.</p>
</div>

<div class="signature">
    <div class="section-title">5. Подписи сторон</div>
    <div class="signature-line">
        <strong>Продавец:</strong> ___________________ / {datetime.datetime.now().strftime('%d.%m.%Y')}
    </div>
    <div class="signature-line">
        <strong>Покупатель:</strong> ___________________ / {datetime.datetime.now().strftime('%d.%m.%Y')}
    </div>
</div>

</body>
</html>"""

        # Возвращаем как HTML который можно открыть в Word
        return Response(
            content=contract_content,
            media_type='application/msword',
            headers={'Content-Disposition': f'attachment; filename="contract_car_{car_id}.doc"'}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/export/acceptance-act/{car_id}")
async def export_acceptance_act(car_id: int):
    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        cur.execute("""
            SELECT c.*, m.name as model_name, man.name as manufacturer_name
            FROM cars c
            JOIN models m ON c.model_id = m.id
            JOIN manufacturers man ON m.manufacturer_id = man.id
            WHERE c.id = %s
        """, (car_id,))
        car = cur.fetchone()

        cur.close()
        conn.close()

        # Создаем PDF с поддержкой кириллицы
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)

        # Используем шрифт поддерживающий кириллицу
        try:
            p.setFont("Arial", 16)
        except:
            p.setFont("Helvetica", 16)  # Fallback

        # Заголовок
        p.drawString(100, 800, "АКТ ПРИЕМА-ПЕРЕДАЧИ АВТОМОБИЛЯ")

        p.setFont("Arial", 10) if 'Arial' in p.getAvailableFonts() else p.setFont("Helvetica", 10)
        p.drawString(100, 780, f"г. Москва {datetime.datetime.now().strftime('%d.%m.%Y')} г.")

        # Данные автомобиля
        p.setFont("Arial", 12) if 'Arial' in p.getAvailableFonts() else p.setFont("Helvetica-Bold", 12)
        p.drawString(100, 750, "1. Продавец передает, а Покупатель принимает автомобиль:")

        p.setFont("Arial", 11) if 'Arial' in p.getAvailableFonts() else p.setFont("Helvetica", 11)
        y_position = 720
        p.drawString(100, y_position, f"Марка, модель: {car['manufacturer_name']} {car['model_name']}")
        p.drawString(100, y_position - 20, f"Год выпуска: {car['year']}")
        p.drawString(100, y_position - 40, f"Цвет: {car['color']}")
        p.drawString(100, y_position - 60, f"Пробег: {car['mileage']} км")
        p.drawString(100, y_position - 80, f"Цена: {car['price']:,.2f} руб.")
        p.drawString(100, y_position - 100, f"VIN: VIN_{car['id']}")
        p.drawString(100, y_position - 120, f"Двигатель №: ДВ_{car['id']}")
        p.drawString(100, y_position - 140, f"Кузов №: КУ_{car['id']}")
        p.drawString(100, y_position - 160, f"Тип двигателя: {car['engine_type']}")
        p.drawString(100, y_position - 180, f"Коробка передач: {car['transmission']}")

        # Условия
        p.setFont("Arial", 12) if 'Arial' in p.getAvailableFonts() else p.setFont("Helvetica-Bold", 12)
        p.drawString(100, 520, "2. Условия передачи:")
        p.setFont("Arial", 11) if 'Arial' in p.getAvailableFonts() else p.setFont("Helvetica", 11)
        p.drawString(100, 500, "Автомобиль осмотрен Покупателем, претензий по качеству нет.")
        p.drawString(100, 480, "Ключи и документы переданы Покупателю.")

        # Подписи
        p.setFont("Arial", 12) if 'Arial' in p.getAvailableFonts() else p.setFont("Helvetica-Bold", 12)
        p.drawString(100, 420, "ПОДПИСИ СТОРОН:")
        p.setFont("Arial", 11) if 'Arial' in p.getAvailableFonts() else p.setFont("Helvetica", 11)
        p.drawString(100, 400, "Продавец: _________________________")
        p.drawString(100, 380, "Покупатель: _________________________")
        p.drawString(100, 360, f"Дата: {datetime.datetime.now().strftime('%d.%m.%Y')}")

        p.save()

        buffer.seek(0)
        return Response(content=buffer.read(), media_type='application/pdf',
                        headers={'Content-Disposition': f'attachment; filename="acceptance_act_{car_id}.pdf"'})

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/export/receipt/{car_id}")
async def export_receipt(car_id: int):
    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        cur.execute("SELECT * FROM cars WHERE id = %s", (car_id,))
        car = cur.fetchone()

        cur.close()
        conn.close()

        # Создаем красивый Excel файл с оформлением как документ
        buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Квитанция"

        # Стили
        title_font = Font(size=14, bold=True)
        header_font = Font(size=12, bold=True)
        normal_font = Font(size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        # Заголовок
        ws.merge_cells('A1:F1')
        ws['A1'] = "КВИТАНЦИЯ"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')

        # Данные квитанции
        data = [
            ["Дата:", datetime.datetime.now().strftime('%d.%m.%Y'), "", "", "", ""],
            ["", "", "", "", "", ""],
            ["Плательщик:", "[ФИО плательщика]", "", "", "", ""],
            ["", "", "", "", "", ""],
            ["Назначение платежа:", "Государственная пошлина за регистрацию ТС", "", "", "", ""],
            ["", "", "", "", "", ""],
            ["Данные автомобиля:", "", "", "", "", ""],
            ["Марка, модель:",
             f"{car['manufacturer_name'] if 'manufacturer_name' in car else 'Автосалон'} {car['model_name'] if 'model_name' in car else 'Данные из системы'}",
             "", "", "", ""],
            ["Год выпуска:", car['year'], "", "", "", ""],
            ["Цвет:", car['color'], "", "", "", ""],
            ["VIN:", f"VIN_{car['id']}", "", "", "", ""],
            ["", "", "", "", "", ""],
            ["Сумма госпошлины:", "3 300 руб.", "", "", "", ""],
            ["", "", "", "", "", ""],
            ["Подпись плательщика:", "___________________", "", "", "", ""],
        ]

        # Заполняем данные
        for row_idx, row_data in enumerate(data, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = normal_font
                cell.border = border
                if col_idx == 1 and value and value.endswith(':'):
                    cell.font = header_font

        # Настраиваем ширину колонок
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15

        wb.save(buffer)
        buffer.seek(0)

        return Response(content=buffer.read(),
                        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        headers={'Content-Disposition': f'attachment; filename="receipt_{car_id}.xlsx"'})

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == '__main__':
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=8000)